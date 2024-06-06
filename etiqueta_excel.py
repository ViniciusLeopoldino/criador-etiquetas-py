import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import landscape
from reportlab.pdfgen import canvas

def create_labels_from_excel(excel_file, output_pdf, label_width, label_height, font_size):
    try:
        # Carregar o arquivo Excel em modo somente leitura
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active

        # Configurar o PDF
        c = canvas.Canvas(output_pdf, pagesize=(label_height, label_width), bottomup=0)  # Páginas na orientação paisagem
        c.setFont("Helvetica", font_size)

        # Iterar sobre as linhas do arquivo Excel
        for row in ws.iter_rows(values_only=True):
            code, description = row[:2]  # Supondo que a primeira coluna é o código e a segunda é a descrição

            # Calcular as coordenadas para centralizar o texto
            text_width = c.stringWidth(f"Código: {code}", "Helvetica", font_size)
            text_height = font_size  # Altura do texto

            # Coordenadas para centralizar horizontalmente
            x = (label_width - text_width) / 2

            # Coordenadas para centralizar verticalmente
            y = (label_height - text_height) / 2

            # Verificar se a altura do texto ultrapassa a altura da etiqueta
            if text_height > label_height:
                messagebox.showwarning("Aviso", "Altura da etiqueta é insuficiente para o texto.")
                return

            # Verificar se o texto cabe na largura da etiqueta
            if text_width > label_width:
                messagebox.showwarning("Aviso", "Largura da etiqueta é insuficiente para o texto.")
                return

            # Desenhar o código
            c.drawString(x, y, f"Código: {code}")

            # Calcular a posição vertical para a descrição
            y_description = y - text_height - 5

            # Verificar se a descrição cabe na altura da etiqueta
            if y_description < 0:
                messagebox.showwarning("Aviso", "Altura da etiqueta é insuficiente para a descrição.")
                return

            # Desenhar a descrição
            c.drawString(x, y_description, f"Descrição: {description}")

            # Adicionar uma nova página se necessário
            c.showPage()

        # Salvar o PDF
        c.save()
        messagebox.showinfo("Concluído", "As etiquetas foram criadas com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao processar o arquivo Excel: {e}")

def select_excel_file():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry_excel.delete(0, tk.END)
    entry_excel.insert(0, filename)

def select_output_folder():
    foldername = filedialog.askdirectory()
    entry_output.delete(0, tk.END)
    entry_output.insert(0, foldername)

def create_labels():
    excel_file = entry_excel.get()
    output_folder = entry_output.get()
    width_text = entry_width.get()
    height_text = entry_height.get()
    font_size_text = entry_font_size.get()

    if not excel_file:
        messagebox.showerror("Erro", "Selecione um arquivo Excel.")
        return

    if not output_folder:
        messagebox.showerror("Erro", "Selecione uma pasta de saída.")
        return

    if not width_text or not height_text or not font_size_text:
        messagebox.showerror("Erro", "Digite valores válidos para largura, altura e tamanho da fonte.")
        return

    try:
        label_width = int(width_text)
        label_height = int(height_text)
        font_size = int(font_size_text)
    except ValueError:
        messagebox.showerror("Erro", "Digite valores numéricos para largura, altura e tamanho da fonte.")
        return

    output_pdf = os.path.join(output_folder, "etiquetas.pdf")
    create_labels_from_excel(excel_file, output_pdf, label_width, label_height, font_size)

# Criar a janela principal
root = tk.Tk()
root.title("Criar Etiquetas em PDF")

# Criar widgets
label_excel = tk.Label(root, text="Selecione o arquivo Excel:")
entry_excel = tk.Entry(root, width=50)
button_excel = tk.Button(root, text="Procurar", command=select_excel_file)

label_output = tk.Label(root, text="Selecione a pasta de saída:")
entry_output = tk.Entry(root, width=50)
button_output = tk.Button(root, text="Procurar", command=select_output_folder)

label_dimensions = tk.Label(root, text="Dimensões da etiqueta (em mm):")
label_width = tk.Label(root, text="Largura:")
entry_width = tk.Entry(root, width=10)
label_height = tk.Label(root, text="Altura:")
entry_height = tk.Entry(root, width=10)
label_font_size = tk.Label(root, text="Tamanho da fonte:")
entry_font_size = tk.Entry(root, width=10)

button_create = tk.Button(root, text="Criar Etiquetas", command=create_labels)

# Layout dos widgets
label_excel.grid(row=0, column=0, padx=10, pady=5, sticky="e")
entry_excel.grid(row=0, column=1, columnspan=2, padx=5, pady=5)
button_excel.grid(row=0, column=3, padx=5, pady=5)

label_output.grid(row=1, column=0, padx=10, pady=5, sticky="e")
entry_output.grid(row=1, column=1, columnspan=2, padx=5, pady=5)
button_output.grid(row=1, column=3, padx=5, pady=5)

label_dimensions.grid(row=2, column=0, padx=10, pady=5, sticky="e")
label_width.grid(row=2, column=1, padx=5, pady=5, sticky="e")
entry_width.grid(row=2, column=2, padx=5, pady=5, sticky="w")
label_height.grid(row=2, column=3, padx=5, pady=5, sticky="e")
entry_height.grid(row=2, column=4, padx=5, pady=5, sticky="w")
label_font_size.grid(row=2, column=5, padx=5, pady=5, sticky="e")
entry_font_size.grid(row=2, column=6, padx=5, pady=5, sticky="w")

button_create.grid(row=3, column=0, columnspan=7, padx=10, pady=10)

# Iniciar a aplicação
root.mainloop()

